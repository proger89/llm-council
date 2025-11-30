"""
File text extraction for LLM context.
Extracts text from various file formats to provide as context to the council.
"""
import os
from pathlib import Path
from typing import Optional
import asyncio

from .config import MAX_CHARS_PER_FILE, MAX_TOTAL_CONTEXT_CHARS

# Text-based extensions that can be read directly
TEXT_EXTENSIONS = {
    ".txt", ".md", ".json", ".xml", ".csv", ".yaml", ".yml",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css", ".scss",
    ".java", ".c", ".cpp", ".h", ".hpp", ".cs", ".go", ".rs", ".rb",
    ".php", ".sql", ".sh", ".bash", ".zsh", ".ps1",
    ".ini", ".cfg", ".conf", ".env", ".toml",
}


def get_file_extension(filename: str) -> str:
    """Get file extension in lowercase."""
    return Path(filename).suffix.lower()


def extract_text_from_text_file(file_path: str, max_chars: int = MAX_CHARS_PER_FILE) -> str:
    """Extract text from a plain text file."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(max_chars + 100)  # Read a bit extra to check if truncated
            
        if len(content) > max_chars:
            content = content[:max_chars] + "\n[... содержимое обрезано ...]"
        
        return content.strip()
    except Exception as e:
        return f"[Ошибка чтения файла: {str(e)}]"


def extract_text_from_pdf(file_path: str, max_chars: int = MAX_CHARS_PER_FILE) -> str:
    """Extract text from a PDF file using pdfplumber."""
    try:
        import pdfplumber
        
        text_parts = []
        total_chars = 0
        
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if total_chars + len(page_text) > max_chars:
                    # Truncate at limit
                    remaining = max_chars - total_chars
                    text_parts.append(page_text[:remaining])
                    text_parts.append("\n[... содержимое обрезано ...]")
                    break
                text_parts.append(page_text)
                total_chars += len(page_text)
        
        return "\n".join(text_parts).strip()
    except ImportError:
        return "[PDF: библиотека pdfplumber не установлена]"
    except Exception as e:
        return f"[Ошибка чтения PDF: {str(e)}]"


def extract_text_from_docx(file_path: str, max_chars: int = MAX_CHARS_PER_FILE) -> str:
    """Extract text from a DOCX file using python-docx."""
    try:
        from docx import Document
        
        doc = Document(file_path)
        text_parts = []
        total_chars = 0
        
        for para in doc.paragraphs:
            para_text = para.text
            if total_chars + len(para_text) > max_chars:
                remaining = max_chars - total_chars
                text_parts.append(para_text[:remaining])
                text_parts.append("\n[... содержимое обрезано ...]")
                break
            text_parts.append(para_text)
            total_chars += len(para_text)
        
        return "\n".join(text_parts).strip()
    except ImportError:
        return "[DOCX: библиотека python-docx не установлена]"
    except Exception as e:
        return f"[Ошибка чтения DOCX: {str(e)}]"


def extract_text_from_xlsx(file_path: str, max_chars: int = MAX_CHARS_PER_FILE) -> str:
    """Extract text from an XLSX file using openpyxl."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        total_chars = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"=== Лист: {sheet_name} ===")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if total_chars + len(row_text) > max_chars:
                    text_parts.append("[... содержимое обрезано ...]")
                    wb.close()
                    return "\n".join(text_parts).strip()
                text_parts.append(row_text)
                total_chars += len(row_text) + 1
        
        wb.close()
        return "\n".join(text_parts).strip()
    except ImportError:
        return "[XLSX: библиотека openpyxl не установлена]"
    except Exception as e:
        return f"[Ошибка чтения XLSX: {str(e)}]"


def extract_text_from_file(file_info: dict) -> str:
    """
    Extract text from a single file based on its type.
    
    Args:
        file_info: Dict with 'filename', 'path', 'mime_type' keys
        
    Returns:
        Extracted text content
    """
    filename = file_info["filename"]
    file_path = file_info["path"]
    ext = get_file_extension(filename)
    
    if ext in TEXT_EXTENSIONS:
        return extract_text_from_text_file(file_path)
    elif ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext in (".docx", ".doc"):
        if ext == ".doc":
            return f"[{filename}: формат .doc не поддерживается, используйте .docx]"
        return extract_text_from_docx(file_path)
    elif ext in (".xlsx", ".xls"):
        if ext == ".xls":
            return f"[{filename}: формат .xls не поддерживается, используйте .xlsx]"
        return extract_text_from_xlsx(file_path)
    elif ext in (".pptx", ".ppt"):
        return f"[{filename}: извлечение текста из PowerPoint пока не реализовано]"
    else:
        return f"[{filename}: неподдерживаемый формат для извлечения текста]"


async def extract_text_from_files(files_info: list[dict]) -> str:
    """
    Extract text from multiple files and format for LLM context.
    
    Args:
        files_info: List of dicts with file information
        
    Returns:
        Formatted context string with all file contents
    """
    if not files_info:
        return ""
    
    lines = ["=== Вложения ==="]
    total_chars = 0
    
    for file_info in files_info:
        filename = file_info["filename"]
        
        # Extract text (run in thread pool to avoid blocking)
        loop = asyncio.get_event_loop()
        content = await loop.run_in_executor(
            None, extract_text_from_file, file_info
        )
        
        # Format file block
        file_block = f"\n[FILE: {filename}]\n{content}\n[/FILE]"
        
        # Check total limit
        if total_chars + len(file_block) > MAX_TOTAL_CONTEXT_CHARS:
            remaining = MAX_TOTAL_CONTEXT_CHARS - total_chars
            if remaining > 100:
                file_block = file_block[:remaining] + "\n[... остальные файлы опущены ...]"
                lines.append(file_block)
            else:
                lines.append(f"\n[FILE: {filename}]\n[... файл опущен из-за лимита символов ...]\n[/FILE]")
            break
        
        lines.append(file_block)
        total_chars += len(file_block)
    
    return "\n".join(lines)

